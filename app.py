import os
import json
from flask import Flask, render_template, request, jsonify, flash, redirect, url_for, session, send_file
from werkzeug.utils import secure_filename
import PyPDF2
import pandas as pd
from datetime import datetime 
import re
from difflib import SequenceMatcher
import uuid
import tabula
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['PERMANENT_SESSION_LIFETIME'] = 1800  # 30 minutes session lifetime

# Ensure upload and data directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('data', exist_ok=True)

ALLOWED_EXTENSIONS = {'pdf', 'csv', 'xlsx'}

def allowed_file(filename):     
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_transactions(transactions, session_id):
    """Save transactions to a file"""
    file_path = os.path.join('data', f'{session_id}_transactions.json')
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(transactions, f, ensure_ascii=False)

def load_transactions(session_id):
    """Load transactions from a file"""
    file_path = os.path.join('data', f'{session_id}_transactions.json')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def delete_transaction_file(session_id):
    """Delete transaction file"""
    file_path = os.path.join('data', f'{session_id}_transactions.json')
    if os.path.exists(file_path):
        os.remove(file_path)

def extract_text_from_pdf(file_path):
    text = ""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text()
        return text
    except Exception as e:
        print(f"Error reading PDF: {str(e)}")
        return ""

def extract_text_from_csv(file_path):
    try:
        # Try different encodings
        encodings = ['utf-8', 'latin1', 'iso-8859-1']
        df = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
                
        if df is None:
            print("Could not read CSV with any encoding")
            return None
            
        # Remove any unnamed columns
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        return df
    except Exception as e:
        print(f"Error reading CSV: {str(e)}")
        return None

def extract_text_from_xlsx(file_path):
    try:
        # Check if openpyxl is available
        try:
            import openpyxl
            print("openpyxl is available")
        except ImportError:
            print("Warning: openpyxl not found. Installing...")
            import subprocess
            subprocess.check_call(['pip', 'install', 'openpyxl'])
            print("openpyxl installed successfully")
            
        # Read all sheets
        print(f"Reading Excel file: {file_path}")
        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names
        
        print(f"Excel file contains {len(sheet_names)} sheets: {sheet_names}")
        
        # If there's only one sheet, use it
        if len(sheet_names) == 1:
            df = pd.read_excel(file_path, sheet_name=sheet_names[0])
            print(f"Using single sheet: {sheet_names[0]}")
        else:
            # Try to find a sheet that looks like a transaction sheet
            transaction_sheet = None
            for sheet in sheet_names:
                # Read a few rows to check if it looks like a transaction sheet
                sample_df = pd.read_excel(file_path, sheet_name=sheet, nrows=5)
                
                # Check if column names suggest transaction data
                columns = [str(col).lower() for col in sample_df.columns]
                transaction_indicators = ['date', 'amount', 'transaction', 'narration', 'debit', 'credit', 'balance']
                
                if any(indicator in ' '.join(columns) for indicator in transaction_indicators):
                    transaction_sheet = sheet
                    print(f"Found transaction sheet: {sheet}")
                    break
            
            # If we found a transaction sheet, use it
            if transaction_sheet:
                df = pd.read_excel(file_path, sheet_name=transaction_sheet)
            else:
                # Otherwise, use the first non-empty sheet
                df = None
                for sheet in sheet_names:
                    temp_df = pd.read_excel(file_path, sheet_name=sheet)
                    if not temp_df.empty:
                        df = temp_df
                        print(f"Using first non-empty sheet: {sheet}")
                        break
        
        if df is None or df.empty:
            print("All sheets are empty")
            return None
        
        # Handle Excel files with header rows
        # Some bank statements have metadata/headers before the actual column names
        # Try to find the actual header row
        potential_header_rows = []
        for i in range(min(20, len(df))):  # Increased from 10 to 20 to catch deeper headers
            row = df.iloc[i]
            row_values = [str(val).lower() for val in row.values if str(val).strip()]
            row_text = ' '.join(row_values)
            
            # Check if this row has potential column header names
            header_indicators = ['date', 'narration', 'description', 'amount', 'debit', 'credit', 'balance', 'transaction', 's.n.', 'tran. id', 'value date']
            
            # For ICICI specific headers
            icici_indicators = ['s.n.', 'tran. id', 'value date', 'transaction date', 'transaction posted date', 'cheque. no./ref. no.', 'transaction remarks']
            
            # Check for multiple indicators in the same row (more likely to be a header)
            indicator_count = sum(1 for indicator in header_indicators if indicator in row_text)
            icici_indicator_count = sum(1 for indicator in icici_indicators if indicator in row_text)
            
            if indicator_count >= 2 or icici_indicator_count >= 2:
                print(f"Potential header row at index {i}: {row_values}")
                potential_header_rows.append(i)
        
        # If we found potential header rows, use the last one (most likely to be the actual header)
        if potential_header_rows:
            header_row = potential_header_rows[-1]
            if header_row > 0:
                print(f"Found header row at index {header_row}, re-reading Excel with this header")
                df = pd.read_excel(file_path, sheet_name=df.name if hasattr(df, 'name') else 0, header=header_row)
                
                # Print the new columns after header adjustment
                print("New columns after header adjustment:", df.columns.tolist())
        
        # Remove any unnamed columns
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        # Remove rows where all cells are the same value (likely headers/footers)
        df = df[~df.apply(lambda row: row.nunique() == 1, axis=1)]
        
        return df
    except Exception as e:
        print(f"Error reading XLSX: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return None

def parse_transactions(text):
    transactions = []
    lines = text.split('\n')
    
    for line in lines:
        # Look for date patterns (DD/MM/YY or DD/MM/YYYY)
        date_match = re.search(r'\d{2}[/-]\d{2}[/-]\d{2,4}', line)
        if not date_match:
            continue
        
        date = date_match.group()
        remaining_text = line[date_match.end():].strip()
        
        # Value Date (second date in line)
        value_date = ''
        value_date_match = re.search(r'\d{2}[/-]\d{2}[/-]\d{2,4}', remaining_text)
        if value_date_match:
            value_date = value_date_match.group()
            remaining_text = remaining_text[:value_date_match.start()] + remaining_text[value_date_match.end():]
        
        # Find all 10+ digit numbers (possible reference numbers)
        ref_numbers = re.findall(r'\b\d{10,}\b', remaining_text)
        cheque_ref = ref_numbers[-1] if ref_numbers else ''
        for ref in ref_numbers:
            remaining_text = re.sub(re.escape(ref), '', remaining_text)

        # Find all amounts in the line
        amounts = re.findall(r'(\d{1,3}(?:,\d{3})*(?:\.\d{2}))', remaining_text)
        withdrawal = ''
        deposit = ''
        closing_balance = ''

        # Check for explicit withdrawal/deposit indicators
        withdrawal_indicators = ['DR', 'W/D', 'WITHDRAWAL', 'DEBIT', 'PAID', 'TRANSFER', 'CHARGES', 'FEE', 'ATM']
        deposit_indicators = ['CR', 'DEPOSIT', 'CREDIT', 'RECEIVED', 'INTEREST', 'REFUND']
        
        # First check for explicit withdrawal indicators
        withdrawal_match = None
        for indicator in withdrawal_indicators:
            pattern = rf'(\d{{1,3}}(?:,\d{{3}})*(?:\.\d{{2}}))\s*{indicator}'
            match = re.search(pattern, remaining_text, re.IGNORECASE)
            if match:
                withdrawal_match = match
                withdrawal = match.group(1).replace(',', '')
                remaining_text = remaining_text.replace(match.group(0), '', 1)
                if match.group(1) in amounts:
                    amounts.remove(match.group(1))
                break
        
        # Then check for explicit deposit indicators
        deposit_match = None
        for indicator in deposit_indicators:
            pattern = rf'(\d{{1,3}}(?:,\d{{3}})*(?:\.\d{{2}}))\s*{indicator}'
            match = re.search(pattern, remaining_text, re.IGNORECASE)
            if match:
                deposit_match = match
                deposit = match.group(1).replace(',', '')
                remaining_text = remaining_text.replace(match.group(0), '', 1)
                if match.group(1) in amounts:
                    amounts.remove(match.group(1))
                break
        
        # If explicit indicators weren't found, try to determine by context
        if not withdrawal and not deposit:
            # Look for keywords in the narration that suggest withdrawal or deposit
            is_likely_withdrawal = any(keyword in remaining_text.upper() for keyword in withdrawal_indicators)
            is_likely_deposit = any(keyword in remaining_text.upper() for keyword in deposit_indicators)
            
            # If we have clear indicators in text, assign the first amount accordingly
            if is_likely_withdrawal and not is_likely_deposit and amounts:
                withdrawal = amounts[0].replace(',', '')
                amounts.pop(0)
            elif is_likely_deposit and not is_likely_withdrawal and amounts:
                deposit = amounts[0].replace(',', '')
                amounts.pop(0)
            # If no clear indicators or conflicting indicators, use position (HDFC: withdrawal, deposit, closing)
            elif len(amounts) == 3:
                withdrawal, deposit, closing_balance = [a.replace(',', '') for a in amounts]
                amounts = []
            elif len(amounts) == 2:
                # If two amounts, typically first is transaction amount, second is balance
                # Try to determine if it's withdrawal or deposit based on context
                if is_likely_withdrawal:
                    withdrawal, closing_balance = [a.replace(',', '') for a in amounts]
                elif is_likely_deposit:
                    deposit, closing_balance = [a.replace(',', '') for a in amounts]
                else:
                    # Default to withdrawal if can't determine
                    withdrawal, closing_balance = [a.replace(',', '') for a in amounts]
                amounts = []
        
        # If we still have amounts left, assign the last one to closing balance
        if amounts and not closing_balance:
            closing_balance = amounts[-1].replace(',', '')
            
        # Clean up narration: remove extra spaces, trailing numbers, and punctuation
        narration = re.sub(r'\s+', ' ', remaining_text).strip()
        narration = re.sub(r'\s\d{1,3}(?:,\d{3})*(?:\.\d{2})?\s*$', '', narration)
        narration = narration.strip(' -|:')  # Remove trailing separators
            
        # Only add if there's actual content
        if narration or cheque_ref or withdrawal or deposit or closing_balance:
            transactions.append({
                'date': date,
                'cheque_ref': cheque_ref,
                'narration': narration,
                'value_date': value_date,
                'withdrawal': withdrawal,
                'deposit': deposit,
                'closing_balance': closing_balance,
                'raw_text': line
            })
    
    return transactions

def analyze_transaction(transaction):
    # Convert amounts to float
    withdrawal = 0.0
    deposit = 0.0
    
    if transaction['withdrawal']:
        try:
            # Handle different decimal formats
            withdrawal_str = transaction['withdrawal'].replace(',', '')
            if '.' in withdrawal_str:
                withdrawal = float(withdrawal_str)
            else:
                withdrawal = float(withdrawal_str) / 100.0  # Convert to decimal if no decimal point
        except ValueError:
            withdrawal = 0.0
            
    if transaction['deposit']:
        try:
            # Handle different decimal formats
            deposit_str = transaction['deposit'].replace(',', '')
            if '.' in deposit_str:
                deposit = float(deposit_str)
            else:
                deposit = float(deposit_str) / 100.0  # Convert to decimal if no decimal point
        except ValueError:
            deposit = 0.0
    
    # Ensure withdrawal and deposit are mutually exclusive
    # If both are non-zero, determine which one is correct based on narration
    if withdrawal > 0 and deposit > 0:
        narration = transaction['narration'].upper()
        withdrawal_indicators = ['DR', 'W/D', 'WITHDRAWAL', 'DEBIT', 'PAID', 'TRANSFER', 'CHARGES', 'FEE', 'ATM']
        deposit_indicators = ['CR', 'DEPOSIT', 'CREDIT', 'RECEIVED', 'INTEREST', 'REFUND']
        
        is_likely_withdrawal = any(keyword in narration for keyword in withdrawal_indicators)
        is_likely_deposit = any(keyword in narration for keyword in deposit_indicators)
        
        # If narration suggests withdrawal, zero out deposit
        if is_likely_withdrawal and not is_likely_deposit:
            deposit = 0.0
        # If narration suggests deposit, zero out withdrawal
        elif is_likely_deposit and not is_likely_withdrawal:
            withdrawal = 0.0
        # If can't determine, keep the larger value and zero out the smaller one
        else:
            if withdrawal >= deposit:
                deposit = 0.0
            else:
                withdrawal = 0.0
    
    return {
        'date': transaction['date'],
        'cheque_ref': transaction['cheque_ref'],
        'narration': transaction['narration'],
        'withdrawal': withdrawal,
        'deposit': deposit,
        'balance': deposit - withdrawal,
        'raw_text': transaction['raw_text']  # Include raw text in analysis
    }

def parse_transactions_from_df(df):
    transactions = []
    
    # Clean column names - convert to lowercase for case-insensitive matching
    clean_columns = {col.lower().strip(): col for col in df.columns}
    
    for _, row in df.iterrows():
        # Get narration from various possible column names
        narration_col = next((clean_columns[col] for col in ['narration', 'particulars', 'description'] 
                             if col in clean_columns), None)
        
        # Get withdrawal amount from various possible column names
        withdrawal_col = next((clean_columns[col] for col in ['withdrawal amt.', 'withdrawal', 'debit', 'debit amount'] 
                              if col in clean_columns), None)
        
        # Get deposit amount from various possible column names
        deposit_col = next((clean_columns[col] for col in ['deposit amt.', 'deposit', 'credit', 'credit amount'] 
                           if col in clean_columns), None)
        
        # Get date from various possible column names
        date_col = next((clean_columns[col] for col in ['date', 'transaction date', 'tran date'] 
                        if col in clean_columns), None)
        
        # Get reference number from various possible column names
        ref_col = next((clean_columns[col] for col in ['chq./ref.no.', 'cheque no', 'reference no', 'ref no'] 
                       if col in clean_columns), None)
        
        # Get value date from various possible column names
        value_date_col = next((clean_columns[col] for col in ['value dt', 'value date'] 
                              if col in clean_columns), None)
        
        # Get closing balance from various possible column names
        balance_col = next((clean_columns[col] for col in ['closing balance', 'balance'] 
                           if col in clean_columns), None)
        
        # Extract values with fallbacks
        narration = str(row.get(narration_col, '')) if narration_col else ''
        withdrawal = str(row.get(withdrawal_col, '')) if withdrawal_col else ''
        deposit = str(row.get(deposit_col, '')) if deposit_col else ''
        date = str(row.get(date_col, '')) if date_col else ''
        cheque_ref = str(row.get(ref_col, '')) if ref_col else ''
        value_date = str(row.get(value_date_col, '')) if value_date_col else ''
        closing_balance = str(row.get(balance_col, '')) if balance_col else ''
        
        # Clean up values
        narration = narration.strip()
        withdrawal = withdrawal.strip().replace('₹', '').replace(',', '')
        deposit = deposit.strip().replace('₹', '').replace(',', '')
        date = date.strip()
        cheque_ref = cheque_ref.strip()
        value_date = value_date.strip()
        closing_balance = closing_balance.strip().replace('₹', '').replace(',', '')
        
        # Ensure withdrawal and deposit are mutually exclusive
        if withdrawal and deposit:
            # Check for indicators in narration
            narration_upper = narration.upper()
            withdrawal_indicators = ['DR', 'W/D', 'WITHDRAWAL', 'DEBIT', 'PAID', 'TRANSFER', 'CHARGES', 'FEE', 'ATM']
            deposit_indicators = ['CR', 'DEPOSIT', 'CREDIT', 'RECEIVED', 'INTEREST', 'REFUND']
            
            is_withdrawal = any(indicator in narration_upper for indicator in withdrawal_indicators)
            is_deposit = any(indicator in narration_upper for indicator in deposit_indicators)
            
            # Clear the incorrect field based on narration indicators
            if is_withdrawal and not is_deposit:
                deposit = ''
            elif is_deposit and not is_withdrawal:
                withdrawal = ''
            # If can't determine from narration, keep the non-zero value
            else:
                try:
                    withdrawal_val = float(withdrawal) if withdrawal else 0
                    deposit_val = float(deposit) if deposit else 0
                    
                    if withdrawal_val > 0 and deposit_val > 0:
                        # Keep the larger value
                        if withdrawal_val >= deposit_val:
                            deposit = ''
                        else:
                            withdrawal = ''
                except ValueError:
                    # If conversion fails, keep both values
                    pass
        
        transaction = {
            'date': date,
            'narration': narration,
            'cheque_ref': cheque_ref,
            'value_date': value_date,
            'withdrawal': withdrawal,
            'deposit': deposit,
            'closing_balance': closing_balance,
            'raw_text': ' | '.join([str(x) for x in row.values if pd.notna(x)]),
            # Add these fields for compatibility with search and other functions
            'date': date,
            'narration': narration
        }
        
        transactions.append(transaction)
    
    return transactions

def extract_transactions_from_pdf_table(file_path):
    try:
        # Check if Java is available
        java_available = False
        try:
            import subprocess
            result = subprocess.run(['java', '-version'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            java_available = result.returncode == 0
        except:
            java_available = False
            print("Java is not available in the system PATH")
        
        if not java_available:
            print("Java is not available, skipping tabula extraction")
            return []
            
        # Try both lattice and stream mode
        dfs = []
        try:
            import tabula
            dfs = tabula.read_pdf(file_path, pages='all', multiple_tables=True, lattice=True)
            if not dfs or len(dfs) == 0:
                dfs = tabula.read_pdf(file_path, pages='all', multiple_tables=True, stream=True)
        except Exception as e:
            print(f"Error using tabula to extract tables: {str(e)}")
            return []
        
        if not dfs or len(dfs) == 0:
            return []
            
        # Combine all tables into one DataFrame
        combined_df = pd.DataFrame()
        for df in dfs:
            if not df.empty:
                # Clean column names
                df.columns = [str(col).strip() for col in df.columns]
                
                # Check if this looks like a transaction table
                # HDFC typically has columns like Date, Narration, Withdrawal, etc.
                lower_cols = [str(col).lower() for col in df.columns]
                transaction_indicators = ['date', 'narration', 'withdrawal', 'deposit', 'chq', 'ref']
                
                if any(indicator in ' '.join(lower_cols) for indicator in transaction_indicators):
                    # If first DataFrame, use it as base
                    if combined_df.empty:
                        combined_df = df
                    else:
                        # Ensure columns match before appending
                        if len(df.columns) == len(combined_df.columns):
                            combined_df = pd.concat([combined_df, df], ignore_index=True)
                        else:
                            # Try to map columns if they don't match exactly
                            df.columns = combined_df.columns[:len(df.columns)]
                            combined_df = pd.concat([combined_df, df], ignore_index=True)
        
        if combined_df.empty:
            return []
            
        # Clean column names for consistency
        # Map common HDFC column variations to standard names
        column_mapping = {
            'date': 'Date',
            'narration': 'Narration',
            'chq./ref.no.': 'Chq./Ref.No.',
            'chq/ref no': 'Chq./Ref.No.',
            'cheque no': 'Chq./Ref.No.',
            'reference no': 'Chq./Ref.No.',
            'value dt': 'Value Dt',
            'value date': 'Value Dt',
            'withdrawal amt.': 'Withdrawal Amt.',
            'withdrawal': 'Withdrawal Amt.',
            'debit': 'Withdrawal Amt.',
            'debit amount': 'Withdrawal Amt.',
            'deposit amt.': 'Deposit Amt.',
            'deposit': 'Deposit Amt.',
            'credit': 'Deposit Amt.',
            'credit amount': 'Deposit Amt.',
            'closing balance': 'Closing Balance',
            'balance': 'Closing Balance'
        }
        
        # Rename columns based on case-insensitive matching
        new_columns = []
        for col in combined_df.columns:
            col_lower = str(col).lower().strip()
            if col_lower in column_mapping:
                new_columns.append(column_mapping[col_lower])
            else:
                new_columns.append(col)
                
        combined_df.columns = new_columns
        
        # Ensure we have the minimum required columns
        required_columns = ['Date', 'Narration']
        if not all(col in combined_df.columns for col in required_columns):
            return []
            
        # Drop rows where Date is missing or invalid
        combined_df = combined_df.dropna(subset=['Date'])
        combined_df = combined_df[combined_df['Date'].astype(str).str.strip() != '']
        
        # Parse transactions using our improved function
        return parse_transactions_from_df(combined_df)
    except Exception as e:
        print(f"Error extracting table from PDF: {str(e)}")
        return []

# Add a new function to extract SBI transactions from PDF using PyPDF2
def extract_sbi_transactions_from_pdf(file_path):
    """Extract SBI transactions from PDF using PyPDF2 when tabula fails"""
    print("Attempting to extract SBI transactions using PyPDF2")
    try:
        # Extract text from PDF
        text = extract_text_from_pdf(file_path)
        if not text:
            return []
            
        # Check if this looks like an SBI statement
        if 'state bank of india' not in text.lower() and 'sbi' not in text.lower():
            print("This doesn't appear to be an SBI statement")
            return []
            
        # Split into lines and process
        lines = text.split('\n')
        transactions = []
        
        # Look for transaction patterns in SBI statements
        date_pattern = r'\d{2}[/-]\d{2}[/-]\d{2,4}'
        amount_pattern = r'(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)'
        
        in_transaction_section = False
        for i, line in enumerate(lines):
            # Check if we've entered the transaction section
            if not in_transaction_section and ('date' in line.lower() and ('description' in line.lower() or 'particulars' in line.lower()) and ('debit' in line.lower() or 'credit' in line.lower())):
                in_transaction_section = True
                continue
                
            if not in_transaction_section:
                continue
                
            # Check if line contains a date (potential transaction)
            date_match = re.search(date_pattern, line)
            if not date_match:
                continue
                
            date = date_match.group()
            remaining_text = line[date_match.end():].strip()
            
            # Extract description/narration
            # In SBI statements, description is usually between date and amounts
            description = remaining_text
            
            # Try to find amounts in the line
            amounts = re.findall(amount_pattern, remaining_text)
            
            debit = ''
            credit = ''
            balance = ''
            
            # Try to determine which amounts are debit, credit and balance
            if len(amounts) >= 3:
                # Typical format: date, description, debit, credit, balance
                debit = amounts[0].replace(',', '') if amounts[0] else ''
                credit = amounts[1].replace(',', '') if amounts[1] else ''
                balance = amounts[2].replace(',', '') if amounts[2] else ''
                
                # Clean up description by removing amounts
                for amt in amounts:
                    description = description.replace(amt, '').strip()
            elif len(amounts) == 2:
                # Could be: date, description, debit/credit, balance
                # Check description for keywords to determine if debit or credit
                if any(word in description.lower() for word in ['withdrawal', 'debit', 'dr', 'paid', 'purchase']):
                    debit = amounts[0].replace(',', '')
                    balance = amounts[1].replace(',', '')
                else:
                    credit = amounts[0].replace(',', '')
                    balance = amounts[1].replace(',', '')
                    
                # Clean up description by removing amounts
                for amt in amounts:
                    description = description.replace(amt, '').strip()
            
            # Clean up description
            description = re.sub(r'\s+', ' ', description).strip()
            description = re.sub(r'[^\w\s\-\.,]', '', description).strip()
            
            # Only add if we have essential data
            if date and (debit or credit):
                transaction = {
                    'date': date,
                    'narration': description,
                    'cheque_ref': '',  # SBI statements might not have this clearly separated
                    'withdrawal': debit,
                    'deposit': credit,
                    'closing_balance': balance,
                    'raw_text': line
                }
                transactions.append(transaction)
        
        print(f"Extracted {len(transactions)} SBI transactions using PyPDF2")
        return transactions
    except Exception as e:
        print(f"Error extracting SBI transactions from PDF: {str(e)}")
        return []

def detect_bank_type(df):
    """Detect bank type from DataFrame columns and content"""
    # Convert column names to lowercase for case-insensitive matching
    columns = [str(col).lower().strip() for col in df.columns]
    column_text = ' '.join(columns)
    
    print("Detecting bank type from columns:", columns)
    
    # Check for SBI indicators - more comprehensive matching (check SBI first)
    sbi_indicators = [
        'particulars', 'description', 'debit', 'credit', 'balance', 'ref no./cheque no',
        'transaction date', 'value date', 'details', 'transaction details', 'narration',
        'txn date', 'posting date', 'transaction description', 'withdrawal amount', 'deposit amount'
    ]
    
    # Check if "SBI" appears in any column value in the first few rows
    sbi_in_data = False
    try:
        for _, row in df.head(5).iterrows():
            row_text = ' '.join([str(val).lower() for val in row.values if pd.notna(val)])
            if 'sbi' in row_text or 'state bank of india' in row_text:
                sbi_in_data = True
                break
    except:
        pass
    
    sbi_matches = sum(1 for indicator in sbi_indicators if any(indicator in col for col in columns))
    if sbi_matches >= 2 or 'sbi' in column_text or sbi_in_data:
        print(f"Detected SBI bank with {sbi_matches} matching indicators")
        return 'SBI'
    
    # Check for HDFC indicators - more comprehensive matching
    hdfc_indicators = [
        'chq./ref.no.', 'narration', 'withdrawal amt.', 'deposit amt.', 'closing balance',
        'chq/ref', 'withdrawal', 'deposit', 'value dt', 'value date'
    ]
    
    hdfc_matches = sum(1 for indicator in hdfc_indicators if any(indicator in col for col in columns))
    if hdfc_matches >= 2 or 'hdfc' in column_text:
        print(f"Detected HDFC bank with {hdfc_matches} matching indicators")
        return 'HDFC'
    
    # Check for ICICI indicators - more comprehensive matching
    icici_indicators = [
        'tran. id', 'transaction remarks', 'withdrawal amt (inr)', 'deposit amt (inr)', 'balance (inr)',
        'txn id', 'transaction id', 'transaction date', 'value date', 'transaction remarks', 
        'transaction posted date'
    ]
    
    icici_matches = sum(1 for indicator in icici_indicators if any(indicator in col for col in columns))
    if icici_matches >= 2 or 'icici' in column_text:
        print(f"Detected ICICI bank with {icici_matches} matching indicators")
        return 'ICICI'
    
    # If we have debit/credit/date columns but couldn't determine specific bank
    generic_indicators = ['date', 'debit', 'credit', 'description', 'narration', 'balance']
    generic_matches = sum(1 for indicator in generic_indicators if any(indicator in col for col in columns))
    
    if generic_matches >= 3:
        print("Detected generic bank statement format")
        return 'GENERIC'
    
    print("Could not detect bank type, returning UNKNOWN")
    return 'UNKNOWN'

def parse_transactions_from_icici(df):
    transactions = []
    
    # Print column information for debugging
    print("ICICI parser: Processing dataframe with columns:", df.columns.tolist())
    
    # Special handling for ICICI statement format
    # Check if this looks like a metadata section before the actual transaction table
    if len(df.columns) <= 2 and any('transaction period' in str(col).lower() for col in df.columns):
        print("Detected ICICI statement metadata format, searching for transaction table")
        
        # Look for the row with column headers
        header_row = None
        for i in range(min(30, len(df))):
            row_values = [str(val).lower() for val in df.iloc[i].values if pd.notna(val)]
            row_text = ' '.join(row_values)
            
            # Check for ICICI transaction headers
            if 's.n.' in row_text and 'tran. id' in row_text and 'transaction remarks' in row_text:
                header_row = i
                print(f"Found ICICI transaction header at row {i}: {row_values}")
                break
        
        if header_row is not None:
            # Extract the actual transaction data
            transaction_data = []
            column_names = []
            
            # Get column names from the header row
            for col in df.columns:
                header_val = str(df.loc[header_row, col]).strip()
                if header_val:
                    column_names.append(header_val)
                    
            print(f"Extracted column names: {column_names}")
            
            # Extract data rows (starting after header row)
            for i in range(header_row + 1, len(df)):
                row = df.iloc[i]
                row_values = [str(val) for val in row.values if pd.notna(val)]
                
                # Skip empty rows or footer rows
                if not row_values or len(row_values) < 3:
                    continue
                    
                # Skip rows that don't start with a number (likely not transaction rows)
                try:
                    int(str(row_values[0]).strip())
                except ValueError:
                    # Check if this might be a continuation of the previous row
                    if transaction_data and len(row_values) >= 1:
                        # Append to the transaction remarks of the previous transaction
                        if len(transaction_data[-1]) >= 7:  # Ensure we have enough columns
                            transaction_data[-1][6] += " " + " ".join(row_values)
                            print(f"Appended continuation data to previous row: {row_values}")
                    continue
                
                # Add the row data
                transaction_data.append(row_values)
            
            print(f"Extracted {len(transaction_data)} transaction rows")
            
            # Map the data to our transaction format
            for row_data in transaction_data:
                try:
                    # Create a dictionary for this transaction
                    transaction = {
                        'sn': row_data[0] if len(row_data) > 0 else '',
                        'tran_id': row_data[1] if len(row_data) > 1 else '',
                        'value_date': row_data[2] if len(row_data) > 2 else '',
                        'transaction_date': row_data[3] if len(row_data) > 3 else '',
                        'transaction_posted_date': row_data[4] if len(row_data) > 4 else '',
                        'cheque_ref': row_data[5] if len(row_data) > 5 else '',
                        'transaction_remarks': row_data[6] if len(row_data) > 6 else '',
                        'withdrawal': row_data[7] if len(row_data) > 7 else '',
                        'deposit': row_data[8] if len(row_data) > 8 else '',
                        'balance': row_data[9] if len(row_data) > 9 else '',
                        'raw_text': ' | '.join(row_data)
                    }
                    
                    # Add these fields for compatibility with search and other functions
                    transaction['date'] = transaction['value_date'] or transaction['transaction_date']
                    transaction['narration'] = transaction['transaction_remarks']
                    
                    transactions.append(transaction)
                except Exception as e:
                    print(f"Error processing ICICI transaction row: {e}")
                    continue
            
            if transactions:
                print(f"Successfully parsed {len(transactions)} ICICI transactions")
                # Print a few sample transactions for debugging
                for i, t in enumerate(transactions[:3]):
                    print(f"Sample transaction {i+1}:")
                    print(f"  Transaction Remarks: {t.get('transaction_remarks', '')}")
                    print(f"  Transaction ID: {t.get('tran_id', '')}")
                    print(f"  Value Date: {t.get('value_date', '')}")
                return transactions
    
    # If the above special handling didn't work, try the regular approach
    # Clean column names and make them case-insensitive
    clean_columns = {col.lower().strip(): col for col in df.columns}
    
    # Define possible column name variations for ICICI
    sn_cols = ['s.n.', 'sn', 'sr no', 'sr. no.', 'sr no.']
    tran_id_cols = ['tran. id', 'transaction id', 'tran id', 'txn id']
    value_date_cols = ['value date', 'val date']
    transaction_date_cols = ['transaction date', 'txn date', 'date']
    posted_date_cols = ['transaction posted date', 'posted date']
    cheque_ref_cols = ['cheque. no./ref. no.', 'cheque no', 'ref no', 'reference no', 'chq no']
    remarks_cols = ['transaction remarks', 'remarks', 'description', 'particulars', 'narration']
    withdrawal_cols = ['withdrawal amt (inr)', 'withdrawal', 'debit', 'debit amount', 'dr amount']
    deposit_cols = ['deposit amt (inr)', 'deposit', 'credit', 'credit amount', 'cr amount']
    balance_cols = ['balance (inr)', 'balance', 'closing balance']
    
    # Find matching columns
    found_cols = {}
    for col_type, possible_cols in [
        ('sn', sn_cols), 
        ('tran_id', tran_id_cols),
        ('value_date', value_date_cols),
        ('transaction_date', transaction_date_cols),
        ('posted_date', posted_date_cols),
        ('cheque_ref', cheque_ref_cols),
        ('remarks', remarks_cols),
        ('withdrawal', withdrawal_cols),
        ('deposit', deposit_cols),
        ('balance', balance_cols)
    ]:
        for col in possible_cols:
            if col in clean_columns:
                found_cols[col_type] = clean_columns[col]
                print(f"Found {col_type} column: {clean_columns[col]}")
                break
    
    for _, row in df.iterrows():
        try:
            # Get values using case-insensitive column matching
            sn = ''
            if 'sn' in found_cols:
                sn = str(row.get(found_cols['sn'], '')).strip()
                
            tran_id = ''
            if 'tran_id' in found_cols:
                tran_id = str(row.get(found_cols['tran_id'], '')).strip()
                
            value_date = ''
            if 'value_date' in found_cols:
                value_date = str(row.get(found_cols['value_date'], '')).strip()
                
            transaction_date = ''
            if 'transaction_date' in found_cols:
                transaction_date = str(row.get(found_cols['transaction_date'], '')).strip()
                
            transaction_posted_date = ''
            if 'posted_date' in found_cols:
                transaction_posted_date = str(row.get(found_cols['posted_date'], '')).strip()
                
            cheque_ref = ''
            if 'cheque_ref' in found_cols:
                cheque_ref = str(row.get(found_cols['cheque_ref'], '')).strip()
                
            transaction_remarks = ''
            if 'remarks' in found_cols:
                transaction_remarks = str(row.get(found_cols['remarks'], '')).strip()
                
            withdrawal = ''
            if 'withdrawal' in found_cols:
                withdrawal_val = row.get(found_cols['withdrawal'], '')
                if pd.notna(withdrawal_val) and withdrawal_val != '':
                    withdrawal = str(withdrawal_val).strip().replace('₹', '').replace(',', '')
            
            deposit = ''
            if 'deposit' in found_cols:
                deposit_val = row.get(found_cols['deposit'], '')
                if pd.notna(deposit_val) and deposit_val != '':
                    deposit = str(deposit_val).strip().replace('₹', '').replace(',', '')
            
            balance = ''
            if 'balance' in found_cols:
                balance_val = row.get(found_cols['balance'], '')
                if pd.notna(balance_val) and balance_val != '':
                    balance = str(balance_val).strip().replace('₹', '').replace(',', '')
            
            # Skip rows that don't have essential data
            if not (transaction_date or value_date) or (not transaction_remarks and not withdrawal and not deposit):
                continue
                
            # Create transaction object
            transaction = {
                'sn': sn,
                'tran_id': tran_id,
                'value_date': value_date,
                'transaction_date': transaction_date,
                'transaction_posted_date': transaction_posted_date,
                'cheque_ref': cheque_ref,
                'transaction_remarks': transaction_remarks,
                'withdrawal': withdrawal,
                'deposit': deposit,
                'balance': balance,
                'raw_text': ' | '.join([str(x) for x in row.values if pd.notna(x)]),
                # Add these fields for compatibility with search and other functions
                'date': transaction_date or value_date,
                'narration': transaction_remarks
            }
            
            transactions.append(transaction)
        except Exception as e:
            print(f"Error processing ICICI row: {e}")
            continue
    
    # If no transactions were found, try using the generic parser as fallback
    if not transactions:
        print("No transactions found with ICICI parser, trying generic parser")
        return parse_transactions_from_df(df)
    
    print(f"Extracted {len(transactions)} transactions with ICICI parser")
    
    # Print a few sample transactions for debugging
    for i, t in enumerate(transactions[:3]):
        print(f"Sample transaction {i+1}:")
        print(f"  Transaction Remarks: {t.get('transaction_remarks', '')}")
        print(f"  Transaction ID: {t.get('tran_id', '')}")
        print(f"  Value Date: {t.get('value_date', '')}")
    
    return transactions

def parse_transactions_from_sbi(df):
    transactions = []
    print("Parsing SBI statement with columns:", df.columns.tolist())
    
    # Convert column names to lowercase for case-insensitive matching
    clean_columns = {col.lower().strip(): col for col in df.columns}
    print("Clean columns:", list(clean_columns.keys()))
    
    # Define possible column name variations for SBI
    date_cols = ['date', 'txn date', 'transaction date', 'value date', 'posting date']
    description_cols = ['description', 'particulars', 'details', 'narration', 'transaction details']
    ref_cols = ['ref no./cheque no', 'ref no', 'cheque no', 'reference no', 'ref no./cheque no.']
    debit_cols = ['debit', 'withdrawal', 'withdrawal amount', 'dr', 'debit amount']
    credit_cols = ['credit', 'deposit', 'deposit amount', 'cr', 'credit amount']
    balance_cols = ['balance', 'closing balance', 'running balance']
    
    # Find matching columns
    date_col = next((clean_columns[col] for col in date_cols if col in clean_columns), None)
    description_col = next((clean_columns[col] for col in description_cols if col in clean_columns), None)
    ref_col = next((clean_columns[col] for col in ref_cols if col in clean_columns), None)
    debit_col = next((clean_columns[col] for col in debit_cols if col in clean_columns), None)
    credit_col = next((clean_columns[col] for col in credit_cols if col in clean_columns), None)
    balance_col = next((clean_columns[col] for col in balance_cols if col in clean_columns), None)
    
    print(f"Found columns - Date: {date_col}, Description: {description_col}, Ref: {ref_col}, Debit: {debit_col}, Credit: {credit_col}, Balance: {balance_col}")
    
    # Check if we have the minimum required columns
    if not date_col or not (description_col or (debit_col and credit_col)):
        print("Missing essential columns for SBI statement")
        return []
    
    for _, row in df.iterrows():
        try:
            # Extract and clean values
            date = str(row.get(date_col, '')).strip() if date_col else ''
            
            # Get description from appropriate column
            details = ''
            if description_col:
                details = str(row.get(description_col, '')).strip()
            
            # Get reference number if available
            ref_no = ''
            if ref_col:
                ref_no = str(row.get(ref_col, '')).strip()
            
            # Get debit amount if available
            debit = '0'
            if debit_col:
                debit_val = row.get(debit_col, '')
                if pd.notna(debit_val) and debit_val != '':
                    debit = str(debit_val).strip().replace(',', '').replace('₹', '').replace('Rs.', '').replace('Rs', '')
            
            # Get credit amount if available
            credit = '0'
            if credit_col:
                credit_val = row.get(credit_col, '')
                if pd.notna(credit_val) and credit_val != '':
                    credit = str(credit_val).strip().replace(',', '').replace('₹', '').replace('Rs.', '').replace('Rs', '')
            
            # Get balance if available
            balance = '0'
            if balance_col:
                balance_val = row.get(balance_col, '')
                if pd.notna(balance_val) and balance_val != '':
                    balance = str(balance_val).strip().replace(',', '').replace('₹', '').replace('Rs.', '').replace('Rs', '')
            
            # Convert to float for comparison, handling empty strings
            try:
                debit_float = float(debit) if debit.strip() else 0
            except ValueError:
                print(f"Invalid debit value: {debit}")
                debit_float = 0
                
            try:
                credit_float = float(credit) if credit.strip() else 0
            except ValueError:
                print(f"Invalid credit value: {credit}")
                credit_float = 0
            
            # Create transaction object
            transaction = {
                'date': date,
                'narration': details,
                'cheque_ref': ref_no,
                'withdrawal': str(debit_float) if debit_float > 0 else '',
                'deposit': str(credit_float) if credit_float > 0 else '',
                'closing_balance': balance,
                'raw_text': ' | '.join([str(x) for x in row.values if pd.notna(x)])
            }
            
            # Only add if we have at least a date or narration and either withdrawal or deposit
            if (date or details) and (transaction['withdrawal'] or transaction['deposit'] or balance):
                transactions.append(transaction)
                
        except Exception as e:
            print(f"Error processing SBI row: {e}")
            continue
    
    print(f"Extracted {len(transactions)} transactions from SBI statement")
    return transactions

def validate_bank_file(bank_type, filename):
    """Validate if the uploaded file matches the selected bank type"""
    if bank_type == 'OTHERS':
        return False, 'File upload is not supported for Others category'
        
    # Check file extension
    if not allowed_file(filename):
        return False, 'Invalid file type. Please upload PDF, CSV, or XLSX files'
        
    return True, ''

@app.route('/')
def index():
    # Generate new session ID and clear old data
    session_id = str(uuid.uuid4())
    session['session_id'] = session_id
    delete_transaction_file(session_id)
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file part'})
    
    file = request.files['file']
    bank_type = request.form.get('bank_type', '')
    
    if not bank_type or bank_type == 'OTHERS':
        return jsonify({'success': False, 'error': 'Please select a valid bank type'})
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No selected file'})
    
    # Validate bank-specific file
    is_valid, error_message = validate_bank_file(bank_type, file.filename)
    if not is_valid:
        return jsonify({'success': False, 'error': error_message})
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        transactions = []
        detected_bank_type = bank_type  # Default to user-selected bank type
        
        try:
            if filename.endswith('.pdf'):
                # Special handling for SBI PDF statements
                if bank_type == 'SBI':
                    print("Processing SBI PDF statement")
                    # First try tabula extraction
                    transactions = extract_transactions_from_pdf_table(file_path)
                    
                    # If tabula fails (likely due to Java issues), try our custom SBI PDF parser
                    if not transactions:
                        print("Tabula extraction failed for SBI PDF, trying custom SBI parser")
                        transactions = extract_sbi_transactions_from_pdf(file_path)
                        
                    # If both methods fail, try generic text extraction
                    if not transactions:
                        print("Custom SBI parser failed, trying generic text extraction")
                        text = extract_text_from_pdf(file_path)
                        transactions = parse_transactions(text)
                else:
                    # For other banks, use standard extraction methods
                    transactions = extract_transactions_from_pdf_table(file_path)
                    if not transactions:
                        text = extract_text_from_pdf(file_path)
                        transactions = parse_transactions(text)
            elif filename.endswith(('.csv', '.xlsx')):
                try:
                    # For ICICI and SBI Excel files, try direct extraction first
                    if filename.endswith('.xlsx') and bank_type in ['ICICI', 'SBI']:
                        print(f"Trying direct Excel extraction for {bank_type} statement")
                        transactions = extract_transactions_from_excel_directly(file_path, bank_type)
                        
                        if transactions:
                            print(f"Successfully extracted {len(transactions)} transactions using direct Excel extraction")
                        else:
                            print(f"Direct Excel extraction failed for {bank_type}, falling back to standard methods")
                    
                    # If direct extraction failed or wasn't attempted, use standard methods
                    if not transactions:
                        if filename.endswith('.csv'):
                            df = extract_text_from_csv(file_path)
                        else:
                            df = extract_text_from_xlsx(file_path)
                        
                        if df is None:
                            if filename.endswith('.xlsx') and bank_type == 'ICICI':
                                return jsonify({
                                    'success': False, 
                                    'error': 'Could not read the ICICI Excel file. Please make sure it\'s a valid Excel file and try again. If the problem persists, try converting to CSV format.'
                                })
                            else:
                                return jsonify({'success': False, 'error': 'Could not extract data from file. Please check the file format.'})
                        
                        # Debugging: Print the DataFrame structure
                        print("File:", filename)
                        print("Selected Bank Type:", bank_type)
                        print("DataFrame columns:", df.columns.tolist())
                        print("DataFrame first few rows:")
                        print(df.head())
                        
                        # Auto-detect bank type to verify the selected one
                        auto_detected_bank_type = detect_bank_type(df)
                        print("Auto-detected bank type:", auto_detected_bank_type)
                        
                        # Special handling for ICICI bank statements
                        if bank_type == 'ICICI':
                            print("Processing ICICI bank statement")
                            
                            # Check if the dataframe has columns that look like ICICI format
                            has_icici_columns = any(col.lower() in ['transaction remarks', 'tran. id', 'value date'] 
                                                 for col in [str(c).lower() for c in df.columns])
                            
                            if not has_icici_columns:
                                print("ICICI statement might have header rows, trying to find proper header")
                                # Try to find the header row
                                header_row = None
                                for i in range(min(15, len(df))):
                                    row = df.iloc[i]
                                    row_values = [str(val).lower() for val in row.values if pd.notna(val)]
                                    row_text = ' '.join(row_values)
                                    
                                    # Check if this row looks like an ICICI header
                                    if ('transaction' in row_text and 'remarks' in row_text) or ('value date' in row_text and 'tran' in row_text):
                                        header_row = i
                                        print(f"Found potential ICICI header row at index {i}: {row_values}")
                                        break
                                
                                # If we found a header row, reread the Excel with this header
                                if header_row is not None:
                                    print(f"Rereading with header at row {header_row}")
                                    if filename.endswith('.csv'):
                                        df = pd.read_csv(file_path, header=header_row)
                                    else:
                                        df = pd.read_excel(file_path, header=header_row)
                                    
                                    print("New columns after header adjustment:", df.columns.tolist())
                            
                            # Now try to parse with the ICICI parser
                            transactions = parse_transactions_from_icici(df)
                            
                            # If ICICI parser fails, try the generic parser
                            if not transactions:
                                print("ICICI parser failed, trying generic parser")
                                transactions = parse_transactions_from_df(df)
                        
                        # Special handling for SBI bank statements
                        elif bank_type == 'SBI':
                            print("Processing SBI bank statement")
                            # Try to preprocess the dataframe for SBI format
                            # Some SBI statements have extra header rows or merged cells
                            
                            # Check if the dataframe has proper column names
                            has_proper_columns = any(col.lower() in ['date', 'description', 'debit', 'credit', 'balance'] 
                                                   for col in [str(c).lower() for c in df.columns])
                            
                            if not has_proper_columns:
                                print("SBI statement might have header rows, trying to find proper header")
                                # Try to find the header row
                                header_row = None
                                for i in range(min(10, len(df))):
                                    row = df.iloc[i]
                                    row_values = [str(val).lower() for val in row.values if pd.notna(val)]
                                    row_text = ' '.join(row_values)
                                    
                                    # Check if this row looks like a header
                                    if ('date' in row_text and ('debit' in row_text or 'credit' in row_text or 'balance' in row_text)):
                                        header_row = i
                                        print(f"Found potential header row at index {i}: {row_values}")
                                        break
                                
                                # If we found a header row, reread the Excel with this header
                                if header_row is not None:
                                    print(f"Rereading with header at row {header_row}")
                                    if filename.endswith('.csv'):
                                        df = pd.read_csv(file_path, header=header_row)
                                    else:
                                        df = pd.read_excel(file_path, header=header_row)
                                    
                                    print("New columns after header adjustment:", df.columns.tolist())
                            
                            # Now try to parse with the SBI parser
                            transactions = parse_transactions_from_sbi(df)
                            
                            # If SBI parser fails, try the generic parser
                            if not transactions:
                                print("SBI parser failed, trying generic parser")
                                transactions = parse_transactions_from_df(df)
                        
                        # Use appropriate parser based on selected bank type
                        elif bank_type == 'HDFC':
                            # Try with the generic parser which works well for HDFC format
                            transactions = parse_transactions_from_df(df)
                        else:
                            # For other bank types, use the generic parser
                            transactions = parse_transactions_from_df(df)
                        
                        # If the selected bank type parser failed but auto-detection worked, try that parser
                        if not transactions and auto_detected_bank_type not in ['UNKNOWN', 'GENERIC', bank_type]:
                            print(f"Trying auto-detected bank type ({auto_detected_bank_type}) parser")
                            if auto_detected_bank_type == 'ICICI':
                                transactions = parse_transactions_from_icici(df)
                            elif auto_detected_bank_type == 'SBI':
                                transactions = parse_transactions_from_sbi(df)
                            
                            # If auto-detection parser worked, update the bank type
                            if transactions:
                                detected_bank_type = auto_detected_bank_type
                                print(f"Using auto-detected bank type: {detected_bank_type}")
                except Exception as e:
                    import traceback
                    print(f"Error processing {bank_type} {filename.split('.')[-1]} file: {str(e)}")
                    print(traceback.format_exc())
                    
                    if bank_type == 'ICICI' and 'openpyxl' in str(e):
                        return jsonify({
                            'success': False, 
                            'error': 'Could not process ICICI Excel file. The openpyxl library is required. Please try again or convert to CSV format.'
                        })
                    else:
                        return jsonify({
                            'success': False, 
                            'error': f'Error processing {bank_type} file: {str(e)}'
                        })
            else:
                return jsonify({'success': False, 'error': 'Invalid file type'})
            
            if not transactions:
                if bank_type == 'ICICI':
                    return jsonify({
                        'success': False, 
                        'error': 'Could not extract transactions from ICICI statement. Please check if the file format is correct and contains transaction data.'
                    })
                elif bank_type == 'SBI':
                    return jsonify({
                        'success': False, 
                        'error': 'Could not extract transactions from SBI statement. Please check if the file format is correct and contains transaction data.'
                    })
                else:
                    return jsonify({
                        'success': False, 
                        'error': 'Could not extract transactions from file. Please check if the file format matches the selected bank type.'
                    })
            
            # Print the number of transactions extracted
            print(f"Extracted {len(transactions)} transactions")
            
            session_id = session.get('session_id')
            if session_id:
                save_transactions(transactions, session_id)
                session['current_file'] = filename
                # Use the user-selected bank type for consistency, unless auto-detection found a different valid type
                session['bank_type'] = bank_type
            
            return jsonify({
                'success': True,
                'message': f'Successfully uploaded {filename}',
                'filename': filename,
                'transaction_count': len(transactions),
                'bank_type': bank_type,
                'detected_bank_type': detected_bank_type
            })
                    
        except Exception as e:
            import traceback
            print("Error processing file:", str(e))
            print(traceback.format_exc())
            
            # Provide more specific error messages for common issues
            error_msg = str(e)
            if 'openpyxl' in error_msg:
                return jsonify({
                    'success': False, 
                    'error': 'Excel file requires the openpyxl library. Please try again or convert to CSV format.'
                })
            elif 'sheet' in error_msg.lower() or 'xls' in error_msg.lower():
                return jsonify({
                    'success': False, 
                    'error': 'Could not read Excel file. Please check if it\'s a valid Excel file or try converting to CSV format.'
                })
            else:
                return jsonify({'success': False, 'error': str(e)})
    
    return jsonify({'success': False, 'error': 'Invalid file type'})

@app.route('/search', methods=['POST'])
def search():
    try:
        data = request.get_json()
        phrase = data.get('phrase', '').strip().lower()
        date_range = data.get('date_range', {})
        
        if not phrase and not (date_range.get('start') or date_range.get('end')):
            return jsonify({'success': False, 'error': 'Please enter a search phrase or select a date range'})
        
        # Load transactions and bank type
        transactions = load_transactions(session.get('session_id'))
        bank_type = session.get('bank_type', 'HDFC')
        
        if not transactions:
            return jsonify({'success': False, 'error': 'No transactions found. Please upload a file first.'})
        
        print(f"Searching in {len(transactions)} transactions with bank type: {bank_type}")
        print(f"Search phrase: '{phrase}'")
        
        # Print first few transactions to debug
        if bank_type == 'ICICI':
            for i, transaction in enumerate(transactions[:3]):
                print(f"Sample transaction {i+1}:")
                print(f"  Transaction Remarks: {transaction.get('transaction_remarks', '')}")
                print(f"  Transaction ID: {transaction.get('tran_id', '')}")
                print(f"  Value Date: {transaction.get('value_date', '')}")
                print(f"  Raw Text: {transaction.get('raw_text', '')[:100]}...")
        
        # Filter by date range if provided
        if date_range and date_range.get('start') and date_range.get('end'):
            start_date = datetime.strptime(date_range['start'], '%Y-%m-%d')
            end_date = datetime.strptime(date_range['end'], '%Y-%m-%d')
            
            filtered_transactions = []
            for transaction in transactions:
                try:
                    # Use value_date for ICICI bank type
                    date_str = None
                    if bank_type == 'ICICI':
                        date_str = transaction.get('value_date')  # Use value_date for filtering
                    else:
                        date_str = transaction.get('date')  # Fallback for other banks
                    
                    if not date_str:
                        continue
                    
                    # Clean up date string
                    date_str = date_str.strip()
                    
                    # Try parsing the date with multiple formats
                    date_formats = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%b/%Y']  # Added '%d/%b/%Y' for '08/Jan/2025'
                    transaction_date = None
                    
                    for fmt in date_formats:
                        try:
                            transaction_date = datetime.strptime(date_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    if transaction_date and start_date <= transaction_date <= end_date:
                        filtered_transactions.append(transaction)
                except Exception as e:
                    print(f"Error parsing date: {e}")
                    continue
            
            transactions = filtered_transactions
            print(f"After date filtering: {len(transactions)} transactions")
        
        # Search for phrase if provided
        results = []
        if phrase:
            print(f"Searching for phrase: '{phrase}'")
            
            # Special handling for ICICI bank statements
            if bank_type == 'ICICI':
                for transaction in transactions:
                    # Get all searchable fields for ICICI
                    transaction_remarks = str(transaction.get('transaction_remarks', '')).lower()
                    tran_id = str(transaction.get('tran_id', '')).lower()
                    cheque_ref = str(transaction.get('cheque_ref', '')).lower()
                    raw_text = str(transaction.get('raw_text', '')).lower()
                    
                    # Print debug info for a few transactions
                    if len(results) < 2:
                        print(f"Checking transaction: {transaction.get('sn', '')}")
                        print(f"  Transaction Remarks: '{transaction_remarks}'")
                        print(f"  Raw Text: '{raw_text[:100]}...'")
                        print(f"  Contains phrase '{phrase}': {phrase in transaction_remarks or phrase in raw_text}")
                    
                    # Check for match in transaction remarks (most important for ICICI)
                    if phrase in transaction_remarks:
                        result = transaction.copy()
                        result['matched_phrase'] = phrase
                        result['matched_field'] = 'transaction_remarks'
                        result['highlight_text'] = transaction_remarks
                        # Preserve all original ICICI columns for display
                        results.append(result)
                        continue
                    
                    # Check in other fields
                    if phrase in tran_id:
                        result = transaction.copy()
                        result['matched_phrase'] = phrase
                        result['matched_field'] = 'tran_id'
                        results.append(result)
                        continue
                        
                    if phrase in cheque_ref:
                        result = transaction.copy()
                        result['matched_phrase'] = phrase
                        result['matched_field'] = 'cheque_ref'
                        results.append(result)
                        continue
                        
                    # Finally check in raw text
                    if phrase in raw_text:
                        result = transaction.copy()
                        result['matched_phrase'] = phrase
                        result['matched_field'] = 'raw_text'
                        result['highlight_text'] = raw_text
                        results.append(result)
                        continue
            elif bank_type == 'SBI':
                for transaction in transactions:
                    searchable_fields = {
                        'narration': transaction.get('narration', ''),
                        'details': transaction.get('details', ''),
                        'cheque_ref': transaction.get('cheque_ref', ''),
                        'raw_text': transaction.get('raw_text', '')
                    }
                    
                    # Combine all searchable fields into one text
                    searchable_text = ' '.join([str(value).lower() for value in searchable_fields.values()])
                    
                    # Check if phrase is in the searchable text
                    if phrase in searchable_text:
                        result = transaction.copy()
                        
                        # Find which field matched
                        matched_fields = []
                        for field, value in searchable_fields.items():
                            if phrase in str(value).lower():
                                matched_fields.append(field)
                        
                        result['matched_phrase'] = phrase
                        result['matched_fields'] = matched_fields
                        result['highlight_text'] = searchable_text
                        results.append(result)
            else:  # HDFC and others
                for transaction in transactions:
                    searchable_fields = {
                        'narration': transaction.get('narration', ''),
                        'cheque_ref': transaction.get('cheque_ref', ''),
                        'raw_text': transaction.get('raw_text', '')
                    }
                    
                    # Combine all searchable fields into one text
                    searchable_text = ' '.join([str(value).lower() for value in searchable_fields.values()])
                    
                    # Check if phrase is in the searchable text
                    if phrase in searchable_text:
                        result = transaction.copy()
                        
                        # Find which field matched
                        matched_fields = []
                        for field, value in searchable_fields.items():
                            if phrase in str(value).lower():
                                matched_fields.append(field)
                        
                        result['matched_phrase'] = phrase
                        result['matched_fields'] = matched_fields
                        result['highlight_text'] = searchable_text
                        results.append(result)
        else:
            # If only date range is provided, return all transactions in range
            results = transactions
        
        print(f"Found {len(results)} matching transactions")
        
        # Sort results by date
        try:
            # Determine date field based on bank type
            if bank_type == 'ICICI':
                results.sort(key=lambda x: datetime.strptime(x.get('value_date') or x.get('transaction_date', '01/01/2000'), '%d/%m/%Y'), reverse=True)
            else:
                results.sort(key=lambda x: datetime.strptime(x.get('date', '01/01/2000'), '%d/%m/%Y'), reverse=True)
        except Exception as e:
            print(f"Error sorting results: {e}")
            # If date sorting fails, return unsorted results
            pass
        
        return jsonify({
            'success': True, 
            'results': results,
            'total_matches': len(results),
            'bank_type': bank_type
        })
    except Exception as e:
        import traceback
        print(f"Search error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)})

@app.route('/clear_session', methods=['POST'])
def clear_session():
    session_id = session.get('session_id')
    if session_id:
        delete_transaction_file(session_id)
    session.clear()
    return jsonify({'success': True, 'message': 'Session cleared successfully'})

@app.route('/export', methods=['POST'])
def export_results():
    try:
        data = request.get_json()
        phrase = data.get('phrase', '').strip().lower()
        date_range = data.get('date_range', {})
        export_format = data.get('format', 'csv').lower()
        
        transactions = load_transactions(session.get('session_id'))
        bank_type = session.get('bank_type', 'HDFC')
    
        if not transactions:
            return jsonify({'success': False, 'error': 'No transactions found'})
            
        print(f"Exporting {len(transactions)} transactions with bank type: {bank_type}")
        
        # Apply date filtering if provided
        if date_range and date_range.get('start') and date_range.get('end'):
            start_date = datetime.strptime(date_range['start'], '%Y-%m-%d')
            end_date = datetime.strptime(date_range['end'], '%Y-%m-%d')
            
            filtered_transactions = []
            for transaction in transactions:
                # Use value_date for ICICI bank type
                date_str = transaction.get('value_date') if bank_type == 'ICICI' else transaction.get('date')
                
                if date_str:
                    date_str = date_str.strip()
                    try:
                        transaction_date = datetime.strptime(date_str, '%d/%b/%Y')  # Adjust as needed
                        if start_date <= transaction_date <= end_date:
                            filtered_transactions.append(transaction)
                    except ValueError:
                        continue
            
            transactions = filtered_transactions
        
        print(f"Exporting {len(transactions)} matching transactions")
        
        # Create DataFrame with appropriate columns based on bank type
        if bank_type == 'ICICI':
            columns = ['S.N.', 'Tran. Id', 'Value Date', 'Transaction Date', 'Transaction Posted Date',
                       'Cheque No./Ref. No.', 'Transaction Remarks', 'Withdrawal Amt (INR)',
                       'Deposit Amt (INR)', 'Balance (INR)']
            
            # Map transaction fields to columns
            df = pd.DataFrame([{
                'S.N.': t.get('sn', '') or str(i+1),
                'Tran. Id': t.get('tran_id', ''),
                'Value Date': t.get('value_date', ''),
                'Transaction Date': t.get('transaction_date', ''),
                'Transaction Posted Date': t.get('transaction_posted_date', ''),
                'Cheque No./Ref. No.': t.get('cheque_ref', ''),
                'Transaction Remarks': t.get('transaction_remarks', ''),
                'Withdrawal Amt (INR)': t.get('withdrawal', ''),
                'Deposit Amt (INR)': t.get('deposit', ''),
                'Balance (INR)': t.get('balance', '')
            } for i, t in enumerate(transactions)])
        elif bank_type == 'HDFC':
            # Define columns for HDFC
            columns = ['Date', 'Narration', 'Chq./Ref.No.', 'Value Dt',
                       'Withdrawal Amt.', 'Deposit Amt.', 'Closing Balance']
            
            # Map transaction fields to columns for HDFC
            df = pd.DataFrame([{
                'Date': t.get('date', ''),
                'Narration': t.get('narration', ''),
                'Chq./Ref.No.': t.get('cheque_ref', ''),
                'Value Dt': t.get('value_date', ''),
                'Withdrawal Amt.': t.get('withdrawal', ''),
                'Deposit Amt.': t.get('deposit', ''),
                'Closing Balance': t.get('closing_balance', '')
            } for t in transactions])
        else:
            # Handle other bank types similarly if needed
            df = pd.DataFrame()  # Initialize df to avoid UnboundLocalError

        # Fill NaN values
        df = df.fillna('')
        
        # Export based on format
        if export_format == 'csv':
            output = BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name='search_results.csv'
            )
        elif export_format == 'excel':
            output = BytesIO()
            df.to_excel(output, index=False, engine='openpyxl')
            output.seek(0)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='search_results.xlsx'
            )
        elif export_format == 'pdf':
            output = BytesIO()
            
            # Create PDF document
            doc = SimpleDocTemplate(
                output,
                pagesize=landscape(A4),  # Ensure landscape for wider tables
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )
            
            # Prepare data for table
            data = [columns]  # Add headers
            for _, row in df.iterrows():
                data.append([str(row[col]) for col in columns])
            
            # Create table
            table = Table(data, colWidths=[1.5 * inch] * len(columns), repeatRows=1)  # Set column widths
            
            # Add style
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ])
            table.setStyle(style)
            
            # Build PDF
            elements = []
            elements.append(table)
            doc.build(elements)
            
            output.seek(0)
            return send_file(
                output,
                mimetype='application/pdf',
                as_attachment=True,
                download_name='search_results.pdf'
            )
        
        return jsonify({'success': False, 'error': 'Invalid export format'})
    except Exception as e:
        import traceback
        print(f"Export error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)})

@app.route('/dashboard')
def dashboard():
    transactions = session.get('transactions', [])
    current_file = session.get('current_file', '')
    bank_type = session.get('bank_type', 'UNKNOWN')
    return render_template('dashboard.html', transactions=transactions, filename=current_file, bank_type=bank_type)

@app.route('/get_all_transactions')
def get_all_transactions():
    session_id = session.get('session_id')
    if not session_id:
        return jsonify({
            'success': False,
            'error': 'No active session. Please upload a file first.'
        })
    
    transactions = load_transactions(session_id)
    current_file = session.get('current_file', '')
    bank_type = session.get('bank_type', 'UNKNOWN')
    
    if not transactions:
        return jsonify({
            'success': False,
            'error': 'No transactions found. Please upload a file first.'
        })
    
    # Sort transactions by date
    transactions.sort(key=lambda x: datetime.strptime(x['date'], '%d/%m/%y'))
    
    return jsonify({
        'success': True,
        'transactions': transactions,
        'filename': current_file,
        'total_transactions': len(transactions),
        'bank_type': bank_type
    })

# Add a new function to directly read Excel files using openpyxl for complex bank statements
def extract_transactions_from_excel_directly(file_path, bank_type):
    try:
        print(f"Directly reading {bank_type} Excel file using openpyxl")
        
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Get the first sheet
        sheet = wb.active
        
        # For ICICI bank statements
        if bank_type == 'ICICI':
            print("Processing ICICI bank statement with direct Excel access")
            transactions = []
            
            # Find the header row
            header_row = None
            header_col_indices = {}
            
            # ICICI specific headers to look for
            icici_headers = ['s.n.', 'tran. id', 'value date', 'transaction date', 
                           'transaction posted date', 'cheque. no./ref. no.', 
                           'transaction remarks', 'withdrawal amt (inr)', 
                           'deposit amt (inr)', 'balance (inr)']
            
            # Search for header row
            for row_idx in range(1, min(30, sheet.max_row + 1)):
                header_matches = 0
                for col_idx in range(1, min(15, sheet.max_column + 1)):
                    cell_value = str(sheet.cell(row=row_idx, column=col_idx).value or '').lower().strip()
                    if any(header.lower() in cell_value for header in icici_headers):
                        header_matches += 1
                
                if header_matches >= 3:  # If we found at least 3 matching headers
                    header_row = row_idx
                    print(f"Found ICICI header row at {header_row}")
                    break
            
            if not header_row:
                print("Could not find ICICI header row")
                return []
            
            # Map column indices to header names
            for col_idx in range(1, min(15, sheet.max_column + 1)):
                cell_value = str(sheet.cell(row=header_row, column=col_idx).value or '').lower().strip()
                
                # Map columns to standardized names
                if 's.n.' in cell_value or cell_value == 'sn':
                    header_col_indices['sn'] = col_idx
                elif 'tran. id' in cell_value or 'transaction id' in cell_value:
                    header_col_indices['tran_id'] = col_idx
                elif 'value date' in cell_value:
                    header_col_indices['value_date'] = col_idx
                elif 'transaction date' in cell_value:
                    header_col_indices['transaction_date'] = col_idx
                elif 'transaction posted date' in cell_value or 'posted date' in cell_value:
                    header_col_indices['transaction_posted_date'] = col_idx
                elif 'cheque. no./ref. no.' in cell_value or 'cheque no' in cell_value or 'ref no' in cell_value:
                    header_col_indices['cheque_ref'] = col_idx
                elif 'transaction remarks' in cell_value or 'remarks' in cell_value or 'narration' in cell_value:
                    header_col_indices['transaction_remarks'] = col_idx
                elif 'withdrawal' in cell_value or 'debit' in cell_value:
                    header_col_indices['withdrawal'] = col_idx
                elif 'deposit' in cell_value or 'credit' in cell_value:
                    header_col_indices['deposit'] = col_idx
                elif 'balance' in cell_value:
                    header_col_indices['balance'] = col_idx
            
            print(f"Mapped column indices: {header_col_indices}")
            
            # Extract transactions (starting after header row)
            current_transaction = None
            
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                # Check if this is a transaction row (should start with a number in S.N. column)
                sn_value = sheet.cell(row=row_idx, column=header_col_indices.get('sn', 1)).value
                
                # Try to parse S.N. as a number
                is_transaction_row = False
                try:
                    if sn_value and float(str(sn_value).strip()):
                        is_transaction_row = True
                except (ValueError, TypeError):
                    is_transaction_row = False
                
                if is_transaction_row:
                    # Start a new transaction
                    if current_transaction:
                        transactions.append(current_transaction)
                    
                    # Create new transaction
                    current_transaction = {
                        'sn': str(sn_value).strip(),
                        'tran_id': str(sheet.cell(row=row_idx, column=header_col_indices.get('tran_id', 0)).value or '').strip(),
                        'value_date': str(sheet.cell(row=row_idx, column=header_col_indices.get('value_date', 0)).value or '').strip(),
                        'transaction_date': str(sheet.cell(row=row_idx, column=header_col_indices.get('transaction_date', 0)).value or '').strip(),
                        'transaction_posted_date': str(sheet.cell(row=row_idx, column=header_col_indices.get('transaction_posted_date', 0)).value or '').strip(),
                        'cheque_ref': str(sheet.cell(row=row_idx, column=header_col_indices.get('cheque_ref', 0)).value or '').strip(),
                        'transaction_remarks': str(sheet.cell(row=row_idx, column=header_col_indices.get('transaction_remarks', 0)).value or '').strip(),
                        'withdrawal': str(sheet.cell(row=row_idx, column=header_col_indices.get('withdrawal', 0)).value or '').strip().replace(',', ''),
                        'deposit': str(sheet.cell(row=row_idx, column=header_col_indices.get('deposit', 0)).value or '').strip().replace(',', ''),
                        'balance': str(sheet.cell(row=row_idx, column=header_col_indices.get('balance', 0)).value or '').strip().replace(',', '')
                    }
                    
                    # Add compatibility fields
                    current_transaction['date'] = current_transaction['value_date'] or current_transaction['transaction_date']
                    current_transaction['narration'] = current_transaction['transaction_remarks']
                    
                    # Create raw text for searching
                    raw_text_parts = []
                    for col_idx in range(1, min(15, sheet.max_column + 1)):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            raw_text_parts.append(str(cell_value))
                    
                    current_transaction['raw_text'] = ' | '.join(raw_text_parts)
                else:
                    # This might be a continuation of transaction remarks
                    if current_transaction:
                        # Check if there's any content in this row
                        row_has_content = False
                        row_content = []
                        
                        for col_idx in range(1, min(15, sheet.max_column + 1)):
                            cell_value = sheet.cell(row=row_idx, column=col_idx).value
                            if cell_value:
                                row_has_content = True
                                row_content.append(str(cell_value))
                        
                        if row_has_content:
                            # Append to transaction remarks
                            current_transaction['transaction_remarks'] += " " + " ".join(row_content)
                            current_transaction['narration'] = current_transaction['transaction_remarks']
                            current_transaction['raw_text'] += " | " + " ".join(row_content)
            
            # Add the last transaction if exists
            if current_transaction:
                transactions.append(current_transaction)
            
            print(f"Extracted {len(transactions)} ICICI transactions using direct Excel access")
            
            # Print a few sample transactions
            for i, t in enumerate(transactions[:3]):
                print(f"Sample transaction {i+1}:")
                print(f"  S.N.: {t.get('sn', '')}")
                print(f"  Transaction Remarks: {t.get('transaction_remarks', '')[:100]}...")
                print(f"  Transaction ID: {t.get('tran_id', '')}")
                print(f"  Value Date: {t.get('value_date', '')}")
            
            return transactions
            
        # For SBI bank statements
        elif bank_type == 'SBI':
            print("Processing SBI bank statement with direct Excel access")
            transactions = []
            
            # Find the header row
            header_row = None
            header_col_indices = {}
            
            # SBI specific headers to look for
            sbi_headers = ['date', 'description', 'ref no./cheque no', 'debit', 'credit', 'balance']
            
            # Search for header row
            for row_idx in range(1, min(30, sheet.max_row + 1)):
                header_matches = 0
                for col_idx in range(1, min(15, sheet.max_column + 1)):
                    cell_value = str(sheet.cell(row=row_idx, column=col_idx).value or '').lower().strip()
                    if any(header.lower() in cell_value for header in sbi_headers):
                        header_matches += 1
                
                if header_matches >= 3:  # If we found at least 3 matching headers
                    header_row = row_idx
                    print(f"Found SBI header row at {header_row}")
                    break
            
            if not header_row:
                print("Could not find SBI header row")
                return []
            
            # Map column indices to header names
            for col_idx in range(1, min(15, sheet.max_column + 1)):
                cell_value = str(sheet.cell(row=header_row, column=col_idx).value or '').lower().strip()
                
                # Map columns to standardized names
                if 'date' in cell_value and not 'value' in cell_value:
                    header_col_indices['date'] = col_idx
                elif 'description' in cell_value or 'particulars' in cell_value or 'narration' in cell_value:
                    header_col_indices['description'] = col_idx
                elif 'ref no' in cell_value or 'cheque no' in cell_value:
                    header_col_indices['ref_no'] = col_idx
                elif 'debit' in cell_value or 'withdrawal' in cell_value:
                    header_col_indices['debit'] = col_idx
                elif 'credit' in cell_value or 'deposit' in cell_value:
                    header_col_indices['credit'] = col_idx
                elif 'balance' in cell_value:
                    header_col_indices['balance'] = col_idx
            
            print(f"Mapped column indices: {header_col_indices}")
            
            # Extract transactions (starting after header row)
            current_transaction = None
            
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                # Check if this is a transaction row (should have a date in date column)
                date_value = None
                if 'date' in header_col_indices:
                    date_value = sheet.cell(row=row_idx, column=header_col_indices['date']).value
                
                # Check if this row has a date or any debit/credit amount
                is_transaction_row = False
                if date_value:
                    is_transaction_row = True
                else:
                    # Check if there's a debit or credit amount
                    debit_value = sheet.cell(row=row_idx, column=header_col_indices.get('debit', 0)).value
                    credit_value = sheet.cell(row=row_idx, column=header_col_indices.get('credit', 0)).value
                    if debit_value or credit_value:
                        is_transaction_row = True
                
                if is_transaction_row:
                    # Start a new transaction
                    if current_transaction:
                        transactions.append(current_transaction)
                    
                    # Create new transaction
                    current_transaction = {
                        'date': str(date_value or '').strip(),
                        'narration': str(sheet.cell(row=row_idx, column=header_col_indices.get('description', 0)).value or '').strip(),
                        'cheque_ref': str(sheet.cell(row=row_idx, column=header_col_indices.get('ref_no', 0)).value or '').strip(),
                        'withdrawal': str(sheet.cell(row=row_idx, column=header_col_indices.get('debit', 0)).value or '').strip().replace(',', ''),
                        'deposit': str(sheet.cell(row=row_idx, column=header_col_indices.get('credit', 0)).value or '').strip().replace(',', ''),
                        'closing_balance': str(sheet.cell(row=row_idx, column=header_col_indices.get('balance', 0)).value or '').strip().replace(',', '')
                    }
                    
                    # Create raw text for searching
                    raw_text_parts = []
                    for col_idx in range(1, min(15, sheet.max_column + 1)):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            raw_text_parts.append(str(cell_value))
                    
                    current_transaction['raw_text'] = ' | '.join(raw_text_parts)
                else:
                    # This might be a continuation of narration
                    if current_transaction:
                        # Check if there's any content in this row
                        row_has_content = False
                        row_content = []
                        
                        for col_idx in range(1, min(15, sheet.max_column + 1)):
                            cell_value = sheet.cell(row=row_idx, column=col_idx).value
                            if cell_value:
                                row_has_content = True
                                row_content.append(str(cell_value))
                        
                        if row_has_content:
                            # Append to narration
                            current_transaction['narration'] += " " + " ".join(row_content)
                            current_transaction['raw_text'] += " | " + " ".join(row_content)
            
            # Add the last transaction if exists
            if current_transaction:
                transactions.append(current_transaction)
            
            print(f"Extracted {len(transactions)} SBI transactions using direct Excel access")
            
            # Print a few sample transactions
            for i, t in enumerate(transactions[:3]):
                print(f"Sample transaction {i+1}:")
                print(f"  Date: {t.get('date', '')}")
                print(f"  Narration: {t.get('narration', '')[:100]}...")
                print(f"  Withdrawal: {t.get('withdrawal', '')}")
                print(f"  Deposit: {t.get('deposit', '')}")
            
            return transactions
        
        return []
    except Exception as e:
        print(f"Error directly extracting from Excel: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return []

if __name__ == '__main__':
    app.run(debug=True)